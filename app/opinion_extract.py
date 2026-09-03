"""修改意见文件提取：Word / 文本 / Excel / 图片（Gemini 识字）。"""
from __future__ import annotations
import asyncio, base64, csv, io, os, re, shutil
from pathlib import Path

from .config import SCRIPTS_DIR, PYEXE, compare_model_profiles
from .llm import chat, LlmError

PYENV = dict(os.environ, PYTHONIOENCODING="utf-8")

WORD_EXT = {".docx", ".wps"}
TEXT_EXT = {".txt", ".md"}
EXCEL_EXT = {".xlsx", ".xlsm", ".xls", ".csv"}
IMAGE_EXT = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".tif", ".tiff", ".bmp"}
ALLOWED_OPINION_EXT = WORD_EXT | TEXT_EXT | EXCEL_EXT | IMAGE_EXT

MAX_IMAGE_BYTES = 12 * 1024 * 1024
OCR_TIMEOUT_S = 120.0

_MIME = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".webp": "image/webp",
    ".gif": "image/gif",
    ".tif": "image/tiff",
    ".tiff": "image/tiff",
    ".bmp": "image/bmp",
}

OCR_PROMPT = (
    "这是申报书修改意见的图片（打印、扫描或手写均可）。请识别图中全部可见文字，原样抄录。"
    "不要翻译，不要总结，不要分析图片，不要输出思考过程。"
    "若完全没有文字，则 OCR 段只写：（图片中未识别到文字）"
    "必须按下述格式输出，OCR 与 END 标记各占一行：\n"
    "<<<OCR>>>\n"
    "（此处只放从图中抄下的原文）\n"
    "<<<END>>>"
)


def ext_of(name: str) -> str:
    s = str(name or "")
    i = s.rfind(".")
    return s[i:].lower() if i >= 0 else ""


def is_opinion_ext(name: str) -> bool:
    return ext_of(name) in ALLOWED_OPINION_EXT


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _rows_to_text(title: str, rows) -> str:
    lines = ["【工作表：" + str(title or "Sheet") + "】"]
    n = 0
    for row in rows or []:
        cells = [_cell_str(c) for c in (row or [])]
        while cells and not cells[-1]:
            cells.pop()
        if not any(cells):
            continue
        lines.append("\t".join(cells))
        n += 1
    if n == 0:
        return ""
    return "\n".join(lines)


def excel_to_text(path: Path) -> str:
    p = Path(path)
    ext = p.suffix.lower()
    if ext == ".csv":
        return _csv_to_text(p)
    if ext in (".xlsx", ".xlsm"):
        return _xlsx_to_text(p)
    if ext == ".xls":
        return _xls_to_text(p)
    raise ValueError("不是 Excel 文件：" + p.name)


def _csv_to_text(path: Path) -> str:
    raw = path.read_bytes()
    text = None
    for enc in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = raw.decode("utf-8", "replace")
    rows = list(csv.reader(io.StringIO(text)))
    out = _rows_to_text(path.stem, rows)
    if not out.strip():
        raise ValueError("CSV 中没有可提取的文字")
    return out


def _xlsx_to_text(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ValueError("服务器未安装 openpyxl，无法读取 .xlsx") from e
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        raise ValueError("无法打开 Excel：" + str(e)[:180]) from e
    parts = []
    try:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            chunk = _rows_to_text(ws.title, rows)
            if chunk:
                parts.append(chunk)
    finally:
        wb.close()
    if not parts:
        raise ValueError("Excel 中没有可提取的文字")
    return "\n\n".join(parts)


def _xls_to_text(path: Path) -> str:
    try:
        import xlrd
    except ImportError as e:
        raise ValueError("服务器未安装 xlrd，无法读取旧版 .xls") from e
    try:
        book = xlrd.open_workbook(str(path))
    except Exception as e:
        raise ValueError("无法打开 .xls：" + str(e)[:180]) from e
    parts = []
    for sheet in book.sheets():
        rows = []
        for r in range(sheet.nrows):
            rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
        chunk = _rows_to_text(sheet.name, rows)
        if chunk:
            parts.append(chunk)
    if not parts:
        raise ValueError("Excel 中没有可提取的文字")
    return "\n\n".join(parts)


def _sniff_image(data: bytes, name: str) -> tuple[str, str]:
    ext = ext_of(name)
    mime = _MIME.get(ext) or ""
    head = data[:16]
    if head.startswith(b"\xff\xd8\xff"):
        return ".jpg", "image/jpeg"
    if head.startswith(b"\x89PNG\r\n\x1a\n"):
        return ".png", "image/png"
    if head.startswith(b"GIF87a") or head.startswith(b"GIF89a"):
        return ".gif", "image/gif"
    if head.startswith(b"RIFF") and data[8:12] == b"WEBP":
        return ".webp", "image/webp"
    if head.startswith(b"BM"):
        return ".bmp", "image/bmp"
    if head[:4] in (b"II*\x00", b"MM\x00*"):
        return ".tif", "image/tiff"
    if mime:
        return ext, mime
    raise ValueError("不是支持的图片格式（jpg / png / webp / gif / tif / bmp）：" + (name or "file"))


def _prepare_image_bytes(data: bytes, mime: str) -> tuple[bytes, str]:
    """必要时转成 JPEG，避免 TIFF/BMP 不被网关接受。"""
    need = mime not in ("image/jpeg", "image/png", "image/webp", "image/gif") or len(data) > 4 * 1024 * 1024
    if not need:
        return data, mime
    try:
        from PIL import Image
    except ImportError:
        return data, mime
    im = Image.open(io.BytesIO(data))
    if im.mode not in ("RGB", "L"):
        im = im.convert("RGB")
    elif im.mode == "L":
        im = im.convert("RGB")
    w, h = im.size
    mx = 4096
    if max(w, h) > mx:
        im.thumbnail((mx, mx))
    buf = io.BytesIO()
    im.save(buf, format="JPEG", quality=85)
    return buf.getvalue(), "image/jpeg"


async def image_to_text(path: Path) -> str:
    p = Path(path)
    data = p.read_bytes()
    if not data:
        raise ValueError("图片为空：" + p.name)
    if len(data) > MAX_IMAGE_BYTES:
        raise ValueError("图片超过 12MB，请压缩后再上传：" + p.name)
    _ext, mime = _sniff_image(data, p.name)
    data, mime = _prepare_image_bytes(data, mime)
    gem = (compare_model_profiles() or {}).get("gemini") or {}
    if not gem.get("ready"):
        raise ValueError("图片意见需要 Gemini 识别文字。请先在模型配置中填好 Gemini 的地址和密钥。")
    b64 = base64.b64encode(data).decode("ascii")
    messages = [{
        "role": "user",
        "content": [
            {"type": "text", "text": OCR_PROMPT},
            {"type": "image_url", "image_url": {"url": "data:" + mime + ";base64," + b64}},
        ],
    }]
    try:
        r = await chat(
            messages,
            json_mode=False,
            timeout_s=OCR_TIMEOUT_S,
            model=gem.get("id"),
            retries=2,
            apply_profile_timeout=False,
        )
    except LlmError as e:
        raise ValueError("Gemini 识字失败（" + p.name + "）：" + str(e)[:240]) from e
    text = _strip_ocr(str((r or {}).get("content") or ""))
    if not text:
        raise ValueError("Gemini 未从图片中提取到文字：" + p.name)
    return text


def _strip_ocr(text: str) -> str:
    s = str(text or "").strip()
    if not s:
        return ""
    for start, end in (("<<<OCR>>>", "<<<END>>>"), ("<<<TEXT>>>", "<<<END>>>")):
        i = s.find(start)
        if i < 0:
            continue
        rest = s[i + len(start):]
        j = rest.find(end)
        body = rest[:j] if j >= 0 else rest
        body = body.strip()
        if body:
            return body
    fence = chr(96) * 3
    if s.startswith(fence):
        s = re.sub(r"^```(?:text|txt|markdown)?\s*", "", s, count=1)
        if s.endswith(fence):
            s = s[: -len(fence)]
        s = s.strip()
    return s


async def _word_to_txt(src: Path, dst: Path) -> None:
    proc = await asyncio.create_subprocess_exec(
        PYEXE, str(SCRIPTS_DIR / "sb_extract.py"), str(src), str(dst),
        stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE, env=PYENV,
    )
    try:
        out, err = await asyncio.wait_for(proc.communicate(), timeout=120)
    except asyncio.TimeoutError:
        proc.kill()
        raise ValueError("Word 提取超时：" + src.name)
    if (proc.returncode or 0) != 0:
        msg = (err or out or b"").decode("utf-8", "replace")[:200]
        raise ValueError("Word 提取失败：" + src.name + (" " + msg if msg else ""))
    if not dst.exists() or dst.stat().st_size == 0:
        raise ValueError("Word 提取结果为空：" + src.name)


async def ensure_txt(src: Path, dst: Path) -> None:
    """把意见/申报书源文件提取为 utf-8 txt。失败抛 ValueError。"""
    src, dst = Path(src), Path(dst)
    ext = src.suffix.lower()
    dst.parent.mkdir(parents=True, exist_ok=True)
    if ext in TEXT_EXT:
        shutil.copyfile(src, dst)
        return
    if ext in EXCEL_EXT:
        text = excel_to_text(src)
        dst.write_text(text, encoding="utf-8")
        return
    if ext in IMAGE_EXT:
        text = await image_to_text(src)
        dst.write_text(text, encoding="utf-8")
        return
    if ext in WORD_EXT:
        await _word_to_txt(src, dst)
        return
    raise ValueError("不支持的文件类型：" + src.name)
