# -*- coding: utf-8 -*-
"""从申报书标注栏（Word 批注 / Excel 批注）抽出修改意见。"""
from __future__ import annotations
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

NO_OPINION_MSG = "未找到修改意见"
INLINE_OP_NAME = "申报书标注意见.txt"
_MARK = re.compile(r"^<<<标注\s+\d+>>>", re.M)

W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
WORD_EXT = {".docx", ".docm", ".wps"}
EXCEL_EXT = {".xlsx", ".xlsm"}


def _local(tag: str) -> str:
    return tag.split("}")[-1] if tag else ""


def _w_id(el) -> str:
    return str(el.get(W + "id") or el.get("id") or "")


def _para_text(pnode) -> str:
    parts = []
    for node in pnode.iter():
        name = _local(node.tag)
        if name == "t":
            parts.append(node.text or "")
        elif name in ("br", "cr"):
            parts.append("\n")
        elif name == "tab":
            parts.append("\t")
    return "".join(parts)


def _word_comments(path: Path) -> list[dict]:
    path = Path(path)
    try:
        with zipfile.ZipFile(path) as z:
            names = z.namelist()
            if "word/comments.xml" not in names:
                return []
            comments_root = ET.fromstring(z.read("word/comments.xml"))
            doc = None
            if "word/document.xml" in names:
                doc = ET.fromstring(z.read("word/document.xml"))
    except Exception:
        return []

    items = []
    by_id = {}
    for c in comments_root:
        if _local(c.tag) != "comment":
            continue
        cid = _w_id(c)
        body = "\n".join(_para_text(p).strip() for p in c.iter(W + "p") if _para_text(p).strip())
        body = re.sub(r"[ \t]+\n", "\n", body).strip()
        if not body:
            continue
        rec = {
            "id": cid,
            "author": str(c.get(W + "author") or "").strip(),
            "text": body,
            "anchor": "",
        }
        by_id[cid] = rec
        items.append(rec)
    if not items or doc is None:
        return items

    active: dict[str, list[str]] = {}
    for el in doc.iter():
        name = _local(el.tag)
        if name == "commentRangeStart":
            cid = _w_id(el)
            if cid in by_id:
                active[cid] = active.get(cid) or []
        elif name == "commentRangeEnd":
            cid = _w_id(el)
            buf = active.pop(cid, None)
            if buf is not None and cid in by_id and not by_id[cid]["anchor"]:
                by_id[cid]["anchor"] = "".join(buf)
        elif name == "t" and el.text and active:
            for buf in active.values():
                buf.append(el.text)
        elif name in ("br", "cr") and active:
            for buf in active.values():
                buf.append("\n")
        elif name == "tab" and active:
            for buf in active.values():
                buf.append("\t")
    for cid, buf in active.items():
        if cid in by_id and not by_id[cid]["anchor"]:
            by_id[cid]["anchor"] = "".join(buf)
    for rec in items:
        rec["anchor"] = re.sub(r"\s+", " ", rec.get("anchor") or "").strip()
    return items


def _excel_comments(path: Path) -> list[dict]:
    path = Path(path)
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    try:
        wb = load_workbook(path, data_only=False, read_only=False, keep_vba=False)
    except Exception:
        return []
    items = []
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    cmt = getattr(cell, "comment", None)
                    if cmt is None:
                        continue
                    body = str(getattr(cmt, "text", "") or "").strip()
                    if not body:
                        continue
                    author = str(getattr(cmt, "author", "") or "").strip()
                    if author and body.startswith(author):
                        rest = body[len(author):].lstrip(":\n")
                        if rest:
                            body = rest
                    val = cell.value
                    if val is None:
                        anchor = ""
                    elif isinstance(val, float) and val == int(val):
                        anchor = str(int(val))
                    else:
                        anchor = str(val).strip()
                    items.append({
                        "id": ws.title + "!" + str(cell.coordinate),
                        "author": author,
                        "text": body,
                        "anchor": re.sub(r"\s+", " ", anchor)[:240],
                    })
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return items


def extract_comment_items(path: Path) -> list[dict]:
    path = Path(path)
    ext = path.suffix.lower()
    if ext in WORD_EXT:
        return _word_comments(path)
    if ext in EXCEL_EXT:
        return _excel_comments(path)
    return []


def format_inline_opinions(items: list[dict]) -> str:
    lines = []
    n = 0
    for rec in items or []:
        body = str(rec.get("text") or "").strip()
        if not body:
            continue
        n += 1
        lines.append("<<<标注 " + str(n) + ">>>")
        lines.append(str(n) + ". " + body)
        anchor = str(rec.get("anchor") or "").strip()
        if anchor:
            lines.append("标注原文：" + anchor)
        lines.append("")
    return "\n".join(lines).strip()


def extract_inline_opinion_text(path: Path) -> tuple[str, int]:
    items = extract_comment_items(path)
    text = format_inline_opinions(items)
    n = len(re.findall(r"^<<<标注\s+\d+>>>", text, re.M)) if text else 0
    return text, n


def split_inline_units(text: str) -> list[str]:
    s = str(text or "").strip()
    if not s or not _MARK.search(s):
        return []
    parts = _MARK.split(s)
    out = []
    for p in parts:
        t = p.strip()
        if t:
            out.append(t)
    return out
