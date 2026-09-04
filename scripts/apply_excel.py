# -*- coding: utf-8 -*-
"""结构化编辑执行器：按 plan JSON 对 Excel 做单元格文本重写，保持源文件格式。
用法: python apply_excel.py <src.xlsx|xlsm|xls> <out> <backup> <plan.json>
策略: 先单格精确/宽松命中，再按提取时的制表符行拼接跨格替换。
.xlsx/.xlsm 用 openpyxl（xlsm 保留 VBA）；.xls 需本机 Excel COM（FileFormat=56）。
"""
from __future__ import annotations
import json
import os
import re
import shutil
import sys

MAX_SPAN = 8
SEP_ROW = "\t"


def cell_str(v):
    if v is None:
        return ""
    if isinstance(v, bool):
        return "True" if v else "False"
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip() if isinstance(v, str) else str(v)


def loose_regex(find):
    compact = re.sub(r"\s+", "", str(find or ""))
    if not compact:
        return None
    return re.compile(r"\s*".join(re.escape(c) for c in compact))


def _trim_row(cells):
    out = list(cells)
    while out and not out[-1]["text"]:
        out.pop()
    return out


def _join_row(cells):
    return SEP_ROW.join(c["text"] for c in cells)


def _set_cell(cell, text):
    cell["text"] = text
    cell["dirty"] = True


def apply_edits_to_sheets(sheets, edits):
    """sheets: [{name, rows:[{r, cells:[{c,text,dirty,orig}]}]}]. 行列均为 1-based。"""
    results = []
    for e in edits or []:
        find = str(e.get("find", "")).replace("\r\n", "\n").replace("\r", "\n").strip()
        rep = str(e.get("replace", "")).replace("\r\n", "\n").replace("\r", "\n")
        if not find:
            results.append({"find": "", "status": "skip"})
            continue
        status = "miss"
        if _hit_cells(sheets, find, rep, loose=False):
            status = "hit"
        elif _hit_cells(sheets, find, rep, loose=True):
            status = "hit-loose"
        elif _hit_rows(sheets, find, rep, loose=False):
            status = "hit-span"
        elif _hit_rows(sheets, find, rep, loose=True):
            status = "hit-span-loose"
        elif _hit_nearby(sheets, find, rep, loose=False):
            status = "hit-span"
        elif _hit_nearby(sheets, find, rep, loose=True):
            status = "hit-span-loose"
        results.append({"find": find[:100], "status": status})
    return results


def _hit_cells(sheets, find, rep, loose):
    rx = loose_regex(find) if loose else None
    if loose and rx is None:
        return False
    for sh in sheets:
        for row in sh["rows"]:
            for cell in row["cells"]:
                t = cell["text"]
                if not t:
                    continue
                if not loose:
                    p = t.find(find)
                    if p >= 0:
                        _set_cell(cell, t[:p] + rep + t[p + len(find):])
                        return True
                else:
                    m = rx.search(t)
                    if m:
                        _set_cell(cell, t[: m.start()] + rep + t[m.end():])
                        return True
    return False


def _apply_joined(cells, joined, start, end, rep):
    """把拼接串 [start,end) 换成 rep，按原格切回。"""
    spans = []
    pos = 0
    for i, cell in enumerate(cells):
        a = pos
        b = pos + len(cell["text"])
        spans.append((i, a, b))
        pos = b
        if i < len(cells) - 1:
            pos += len(SEP_ROW)
    first = last = None
    for i, a, b in spans:
        if b <= start or a >= end:
            continue
        if first is None:
            first = (i, a, b)
        last = (i, a, b)
    if first is None:
        return False
    fi, fa, fb = first
    li, la, lb = last
    prefix = cells[fi]["text"][: max(0, start - fa)]
    suffix = cells[li]["text"][max(0, end - la):]
    if fi == li:
        _set_cell(cells[fi], prefix + rep + suffix)
        return True
    _set_cell(cells[fi], prefix + rep)
    for i in range(fi + 1, li):
        if cells[i]["text"]:
            _set_cell(cells[i], "")
    _set_cell(cells[li], suffix)
    return True


def _hit_rows(sheets, find, rep, loose):
    rx = loose_regex(find) if loose else None
    if loose and rx is None:
        return False
    for sh in sheets:
        for row in sh["rows"]:
            cells = row["cells"]
            if not cells:
                continue
            joined = _join_row(cells)
            if not loose:
                p = joined.find(find)
                if p >= 0:
                    return _apply_joined(cells, joined, p, p + len(find), rep)
            else:
                m = rx.search(joined)
                if m:
                    return _apply_joined(cells, joined, m.start(), m.end(), rep)
    return False


def _nearby_groups(row_cells, max_w):
    filled = [c for c in row_cells if c["text"]]
    n = len(filled)
    max_w = min(max_w, n)
    for width in range(2, max_w + 1):
        for s in range(0, n - width + 1):
            yield filled[s : s + width]


def _hit_nearby(sheets, find, rep, loose):
    rx = loose_regex(find) if loose else None
    if loose and rx is None:
        return False
    for sh in sheets:
        for row in sh["rows"]:
            for group in _nearby_groups(row["cells"], MAX_SPAN):
                joined = SEP_ROW.join(c["text"] for c in group)
                if not loose:
                    p = joined.find(find)
                    if p >= 0:
                        return _apply_joined(group, joined, p, p + len(find), rep)
                else:
                    m = rx.search(joined)
                    if m:
                        return _apply_joined(group, joined, m.start(), m.end(), rep)
    return False


def load_openpyxl_sheets(path, keep_vba=False):
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=False, keep_vba=keep_vba)
    sheets = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows():
            raw = []
            any_text = False
            r = None
            for cell in row:
                r = cell.row
                val = cell.value
                t = cell_str(val)
                raw.append({"c": cell.column, "text": t, "dirty": False, "orig": val})
                if t:
                    any_text = True
            if not any_text or r is None:
                continue
            cells = _trim_row(raw)
            if cells:
                rows.append({"r": r, "cells": cells})
        sheets.append({"name": ws.title, "rows": rows, "ws": ws})
    return wb, sheets


def write_openpyxl(wb, sheets, out):
    for sh in sheets:
        ws = sh["ws"]
        for row in sh["rows"]:
            for cell in row["cells"]:
                if not cell.get("dirty"):
                    continue
                target = ws.cell(row["r"], cell["c"])
                new = cell["text"]
                orig = cell.get("orig")
                if new == "":
                    target.value = None
                elif isinstance(orig, (int, float)) and not isinstance(orig, bool):
                    try:
                        if "." in new:
                            target.value = float(new)
                        else:
                            target.value = int(new)
                    except ValueError:
                        target.value = new
                else:
                    target.value = new
    wb.save(out)
    wb.close()


def load_xls_sheets(path):
    import xlrd

    book = xlrd.open_workbook(str(path))
    sheets = []
    for si, sheet in enumerate(book.sheets()):
        rows = []
        for r in range(sheet.nrows):
            raw = []
            any_text = False
            for c in range(sheet.ncols):
                t = cell_str(sheet.cell_value(r, c))
                raw.append({"c": c + 1, "text": t, "dirty": False, "orig": sheet.cell_value(r, c)})
                if t:
                    any_text = True
            if not any_text:
                continue
            cells = _trim_row(raw)
            if cells:
                rows.append({"r": r + 1, "cells": cells})
        sheets.append({"name": sheet.name, "rows": rows, "index": si})
    return sheets


def write_xls_com(src, out, sheets):
    try:
        import pythoncom
        import win32com.client
    except ImportError as e:
        raise RuntimeError("写入 .xls 需要本机安装 Microsoft Excel，以及 pywin32") from e

    mutations = []
    for sh in sheets:
        for row in sh["rows"]:
            for cell in row["cells"]:
                if cell.get("dirty"):
                    mutations.append((sh["index"] + 1, row["r"], cell["c"], cell["text"]))
    if not mutations:
        if os.path.abspath(src) != os.path.abspath(out):
            shutil.copyfile(src, out)
        return

    pythoncom.CoInitialize()
    excel = None
    wb = None
    src_abs = str(os.path.abspath(out if os.path.exists(out) else src))
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(src_abs, UpdateLinks=0, ReadOnly=False, AddToMru=False)
        for si, r, c, text in mutations:
            ws = wb.Worksheets(si)
            ws.Cells(r, c).Value = text if text != "" else None
        # 56 = xlExcel8 (.xls)
        out_abs = str(os.path.abspath(out))
        if os.path.abspath(src_abs) == os.path.abspath(out_abs):
            wb.Save()
        else:
            if os.path.exists(out_abs):
                os.remove(out_abs)
            wb.SaveAs(out_abs, FileFormat=56)
        wb.Close(False)
        wb = None
    except Exception as e:
        raise RuntimeError("Excel COM 写入 .xls 失败：" + str(e)[:220]) from e
    finally:
        if wb is not None:
            try:
                wb.Close(False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def apply_file(src, out, backup, edits):
    src, out, backup = str(src), str(out), str(backup)
    ext = os.path.splitext(src)[1].lower()
    shutil.copyfile(src, backup)
    if os.path.abspath(src) != os.path.abspath(out):
        shutil.copyfile(src, out)

    if ext in (".xlsx", ".xlsm"):
        wb, sheets = load_openpyxl_sheets(out, keep_vba=(ext == ".xlsm"))
        try:
            results = apply_edits_to_sheets(sheets, edits)
            write_openpyxl(wb, sheets, out)
        except Exception:
            try:
                wb.close()
            except Exception:
                pass
            raise
        return results

    if ext == ".xls":
        sheets = load_xls_sheets(src)
        results = apply_edits_to_sheets(sheets, edits)
        write_xls_com(src, out, sheets)
        return results

    raise ValueError("不是支持的 Excel 申报书：" + os.path.basename(src))


def main(argv):
    if len(argv) < 5:
        print("usage: apply_excel.py <src.xlsx|xlsm|xls> <out> <backup> <plan.json>", file=sys.stderr)
        return 2
    src, out, backup, plan_path = argv[1], argv[2], argv[3], argv[4]
    plan = json.load(open(plan_path, encoding="utf-8"))
    results = apply_file(src, out, backup, plan.get("edits") or [])
    print(json.dumps({"results": results}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
